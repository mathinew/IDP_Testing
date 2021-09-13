import openpyxl
from selenium import webdriver
import time



def Excel_write():
    path =  "D:\\Workspace\\Project_rbt\\GRM Scenarios.xlsx"
    test = openpyxl.load_workbook(path)
    sheet = test.active
    # print(sheet.cell(row = 2,column = 1).value)
    data = sheet.cell(row = 2,column = 1)
    return (data.value)

    
def sample():
    print("print entry from python file testing")
    







   






    