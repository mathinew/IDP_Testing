import openpyxl

class testing():

   
# generator
# wb = openpyxl.load_workbook(path)
# ws = wb['Commodity Geography']
# rows =ws.iter_rows(min_row =1 , max_row=2, max_col = 4)
# print(rows)
#
# for a,b,c,d in rows:
    # print(a.value , b.value,c.value,d.value )
    
# Normal for loop
    def write_data (self):
        path =  "D:\\Workspace\\Project_rbt\\Data.xlsx"
        wb=openpyxl.load_workbook(path)
        ws = wb['Commodity Geography']
        rows =ws.max_row
        cols = ws.max_column
        
        for r in range(2, rows+1):
                for c in range(2, cols+1):
                    c_value = ws.cell(row=r,column = c).value
                    yield(c_value)
            
            
        
obj = testing()        
a = obj.write_data()

print(int(next(a)))
print("printed")

print(float(next(a)))
print("printed2")

print(int(next(a)))
print("printed3")

print(float(next(a)))
print("printed2")
    
# print(next(a))     
# print("tseting")
# print(next(a))
# print("sample")
# print(next(a))
# print("last line")
# print(next(a)) 
# print("error")
# print(next(a))


       
    
